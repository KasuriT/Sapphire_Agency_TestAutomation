import openpyxl

wb1 = openpyxl.load_workbook("C://Users//lenovo//OneDrive//Documents//CampaignTestData.xlsx")

def fetch_number_of_rows(Sheetname):
    sheet = wb1[Sheetname]
    return sheet.max_row

def fetch_cell_data(Sheetname, row, cell):
    sheet = wb1[Sheetname]
    return sheet.cell(int(row), int(cell)).value