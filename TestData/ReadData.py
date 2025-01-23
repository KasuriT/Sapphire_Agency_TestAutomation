import openpyxl

wb = openpyxl.load_workbook("C://Users//lenovo//OneDrive//Documents//SSLoginTestData.xlsx")

def fetch_number_of_rows(Sheetname):
    sheet = wb[Sheetname]
    return sheet.max_row

def fetch_cell_data(Sheetname, row, cell):
    sheet = wb[Sheetname]
    return sheet.cell(int(row), int(cell)).value